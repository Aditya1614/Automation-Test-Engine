import openpyxl
import io
from typing import List, Dict

class TestCaseParser:
    def __init__(self, file_path: str = None):
        self.file_path = file_path

    def parse(self) -> List[Dict]:
        """
        Parses the Excel file from file path and returns a list of dictionaries representing each test case.
        """
        if not self.file_path:
            raise ValueError("File path not provided.")
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        return self._parse_workbook(wb)
        
    def parse_from_bytes(self, file_bytes: bytes) -> List[Dict]:
        """
        Parses the Excel file from bytes and returns a list of dictionaries.
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        return self._parse_workbook(wb)

    def _parse_workbook(self, wb) -> List[Dict]:
        sheet = wb.active

        headers = []
        test_cases = []
        
        # Read headers
        for cell in sheet[1]:
            headers.append(cell.value.strip() if cell.value else "")

        # Column mapping based on stakeholder requirements
        col_map = {
            "Test Case ID": "id",
            "Positive/Negative/Edge": "type",
            "Positive/Negative": "type",
            "Test Case Scenario": "scenario",
            "Pre-Conditions": "pre_conditions",
            "Test Steps": "test_steps",
            "Test Step Detail": "test_step_detail",
            "Test Data": "test_data",
            "Expected Results": "expected_results",
            "Actual Results": "actual_results",
            "PASSED": "passed",
            "FAILED": "failed",
            "Dokumentasi": "dokumentasi"
        }

        # Find header indices
        header_indices = {}
        for excel_col, internal_key in col_map.items():
            try:
                # If we already mapped this internal key (e.g. from an alternative column name), skip
                if internal_key in header_indices:
                    continue
                idx = headers.index(excel_col)
                header_indices[internal_key] = idx
            except ValueError:
                pass

        id_idx = header_indices.get("id")

        # Read data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or id_idx is None or not row[id_idx]: # Skip empty rows if Test Case ID is empty
                continue
                
            test_case = {}
            for internal_key, idx in header_indices.items():
                val = row[idx]
                test_case[internal_key] = str(val).strip() if val is not None else ""
                
            test_cases.append(test_case)

        return test_cases
